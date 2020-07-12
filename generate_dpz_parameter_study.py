import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
import os
import xml.etree.ElementTree as et

from calibrateGQH import calibrateGQH_Darendeli_LS, calibrateGQH_Menq_LS, calibrateGQH_VuceticDobry_LS, calibrateGQH_Zhang2005_LS
from os_utils import ensure_dir
from xml_indent import indent
from zipfile import ZipFile

def calc_list_numbers(total_num_cases):
    def forward_product(cur_step):
        product = 1
        try:
            for value in total_num_cases[cur_step:]:
                product *= value
        except:
            pass
        return product

    def backward_product(cur_step):
        product = 1
        try:
            for value in total_num_cases[:cur_step+1]:
                product *= value
        except:
            pass
        return product

    forward_list = []
    backward_list = []
    for idx, num in enumerate(total_num_cases):
        forward_list.append(forward_product(idx+1))
        backward_list.append(backward_product(idx-1))
    return forward_list, backward_list

def generate_parameter_study(filename, out_dir = 'parameter_study'):
    def read_Excel_file(xl_filename):
        wb = xl.load_workbook(xl_filename, data_only=True)
        sh = wb['Parameter-Study']

        parameters = []

        # read the parameters
        num_param = 0
        num_cases_list = []
        while True:
            if sh[f"A{num_param+2}"].value == None:
                break
            else:
                cur_param = {}
                cur_param['layer'] = str(sh[f"A{num_param+2}"].value)
                cur_param['variable'] = str(sh[f"B{num_param+2}"].value)
                cur_param['distribution'] = str(sh[f"C{num_param+2}"].value)
                cur_param['min'] = float(sh[f"D{num_param+2}"].value)
                cur_param['max'] = float(sh[f"E{num_param+2}"].value)
                num_cases = int(sh[f"F{num_param+2}"].value)
                cur_param['num_cases'] = num_cases

                num_cases_list.append(num_cases)
                parameters.append(cur_param)
                num_param += 1

        return parameters, num_cases_list
    
    def generate_parametric_list(parameters, num_cases_list):
        foward_list, backward_list = calc_list_numbers(num_cases_list)
        param_list = [{} for ii in range(num_cases_list[0] * foward_list[0])]
        step_count = 1
        for param_ii, param in enumerate(parameters):
            if param['distribution'] == 'Uniform':
                values = np.linspace(param['min'], param['max'], param['num_cases'])
            else:
                raise(f"Distribution {param['distribution']} is not implemented.")
            for ii in range(backward_list[param_ii]):
                for val_ii, val in enumerate(values):
                    for jj in range(foward_list[param_ii]):
                        idx = ii * foward_list[param_ii] * len(values) + val_ii * foward_list[param_ii] + jj
                        if not param['layer'] in param_list[idx]:
                            param_list[idx][param['layer']] = {}
                        param_list[idx][param['layer']][param['variable']] = val
        return param_list

    parameters, num_cases_list = read_Excel_file(filename)
    parameters_list = generate_parametric_list(parameters, num_cases_list)
    ensure_dir(out_dir)
    with open(os.path.join(out_dir,filename[:-5]+'_parameters.txt'), 'w') as f:
        for item in parameters_list:
            f.write(str(item) + '\n')
    wb = xl.load_workbook(filename, data_only=True)

    # read idealized layering boundaries
    sh = wb['Deepsoil-Input']
    layers_bottom = []
    Vs1_list = []
    n_list = []
    Ko_list = []
    ocr_list = []
    pi_list = []
    gamma_list = []
    thickness_list = []
    name_list = []
    Vs_list = []
    layer_bottom = 0.0
    row_num = 2
    while not sh[f'A{row_num}'].value is None:
       thickness = float(sh[f'C{row_num}'].value)
       layer_bottom += thickness
       layers_bottom.append(layer_bottom)
       Vs1_list.append(sh[f'M{row_num}'].value)
       n_list.append(sh[f'P{row_num}'].value)
       Ko_list.append(sh[f'H{row_num}'].value)
       ocr_list.append(sh[f'I{row_num}'].value)
       pi_list.append(sh[f'J{row_num}'].value)
       gamma_list.append(sh[f'E{row_num}'].value)
       thickness_list.append(sh[f'C{row_num}'].value)
       name_list.append(sh[f'B{row_num}'].value)
       Vs_list.append(sh[f'M{row_num}'].value)
       row_num += 1

    sh = wb['Info']
    rock = {}
    rock['unit_weight'] = float(sh['D10'].value)
    rock['Vs'] = float(sh['D9'].value)
    rock['damping'] = float(sh['D11'].value)
    units = str(sh['D16'].value)
    
    # read current calibration
    sh = wb['Deepsoil-Calibrated']
    elements = []
    elem_bottom = 0.0
    row_num = 2
    water_table = -1
    while not sh[f'A{row_num}'].value is None:
       cur_elem = {}
       cur_elem['Name'] = str(sh[f'B{row_num}'].value)
       cur_elem['thickness'] = float(sh[f'C{row_num}'].value)
       elem_bottom += cur_elem['thickness']
       cur_elem['bottom'] = elem_bottom
       cur_elem['water'] = str(sh[f'D{row_num}'].value)
       if water_table == -1 and cur_elem['water'] == 'Y':
           water_table = int(sh[f'A{row_num}'].value)
       cur_elem['unit_weight'] = float(sh[f'E{row_num}'].value)
       cur_elem['Vs'] = float(sh[f'F{row_num}'].value)
       cur_elem['strength'] = float(sh[f'G{row_num}'].value)
       cur_elem['D'] = float(sh[f'H{row_num}'].value)
       cur_elem['t1'] = float(sh[f'I{row_num}'].value)
       cur_elem['t2'] = float(sh[f'J{row_num}'].value)
       cur_elem['t3'] = float(sh[f'K{row_num}'].value)
       cur_elem['t4'] = float(sh[f'L{row_num}'].value)
       cur_elem['t5'] = float(sh[f'M{row_num}'].value)
       cur_elem['p1'] = float(sh[f'N{row_num}'].value)
       cur_elem['p2'] = float(sh[f'O{row_num}'].value)
       cur_elem['p3'] = float(sh[f'P{row_num}'].value)
       cur_elem['Tot_Stress'] = float(sh[f'Q{row_num}'].value)
       cur_elem['PWP'] = float(sh[f'R{row_num}'].value)
       cur_elem['Eff_Stress'] = float(sh[f'S{row_num}'].value)
       cur_elem['Gmax'] = float(sh[f'T{row_num}'].value)

       elements.append(cur_elem)

       row_num += 1

    # write files to be zipped

    # wtite the 'RandomizationSettings' file
    doc = et.Element('doc')
    info = et.SubElement(doc, 'info')
    et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
    et.SubElement(info, 'version').text = "7.0.26.0"
    values = et.SubElement(doc, 'values')
    et.SubElement(values, 'b').text = str(0)
    et.SubElement(values, 'C1').text = str(10.86)
    et.SubElement(values, 'C2').text = str(0.89)
    et.SubElement(values, 'C3').text = str(1.98)
    et.SubElement(values, 'Delta').text = str(13.1)
    et.SubElement(values, 'h0').text = str(0.095)
    et.SubElement(values, 'isThicknessRandomized').text = str(True)
    et.SubElement(values, 'isVelocityRandomized').text = str(True)
    et.SubElement(values, 'isDynamicCurveRandomized').text = str(False)
    et.SubElement(values, 'max_frequency').text = str(25)
    et.SubElement(values, 'max_thickness').text = str(3)
    et.SubElement(values, 'min_thickness').text = str(0.5)
    et.SubElement(values, 'NumberOfRealizations').text = str(len(parameters_list))
    et.SubElement(values, 'Rho0').text = str(0.96)
    et.SubElement(values, 'Rho200').text = str(0.96)
    et.SubElement(values, 'sigma').text = str(0.46)
    et.SubElement(values, 'rho1').text = str(-1)
    et.SubElement(values, 'rho2').text = str(0)
    et.SubElement(values, 'STDEVbound').text = str(1.5)
    et.SubElement(values, 'IsWaterTableChecked').text = str(True)
    et.SubElement(values, 'VRandomMethod').text = str("Geomatrix_AB")
    mydata = et.ElementTree(doc)
    indent(doc)
    with open(os.path.join(out_dir, 'RandomizationSettings'), 'wb') as f:
        mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write profile and meta profile 1
    with open(os.path.join(out_dir, f'Profile1'), 'w') as f:
        f.write("[FILE_VERSION]:[1]\n")
        f.write("[ANALYSIS_DOMAIN]:[TIME+FREQUENCY]\n")
        f.write("[MAX_ITERATIONS]:[15] [COMPLEX_MOD]:[SHAKE_FI] [EFFECTIVE_SSR]:[0.65]\n")
        f.write("[ANALYSIS_TYPE]:[NONLINEAR]\n")
        f.write("[SHEAR_TYPE]:[VELOCITY]\n")
        f.write("[MAX_ITERATIONS]:[5]\n")
        f.write("[ERROR_TOL]:[1E-05]\n")
        f.write("[STEP_CONTROL]:[FLEXIBLE] [MAX_STRAIN_INC]:[0] [INTERPOLATION]:[LINEAR]\n")
        f.write("[VISCOUS_DAMPING]:[FREQ_IND]\n")
        f.write("[DAMPING_UPDATE]:[FALSE]\n")
        f.write(f"[NUM_LAYERS]:[{len(name_list)}]\n")
        f.write("[WATER_TABLE]:[1]\n")

        for layer_ii, layer in enumerate(name_list):
            f.write(f"[LAYER]:[{layer_ii+1}]\n")
            f.write(f"\t[THICKNESS]:[{thickness_list[layer_ii]}] [WEIGHT]:[{gamma_list[layer_ii]}] [SHEAR]:[{Vs_list[layer_ii]}] [SS_DAMP]:[0.02]\n")
            f.write(f"\t[MODEL]:[GQ] [STRENGTH]:[5] [THETA1]:[-20] [THETA2]:[-20] [THETA3]:[0] [THETA4]:[1] [THETA5]:[0.1] [A]:[0]\n")
            f.write(f"\t[MRDF]:[UIUC] [P1]:[1] [P2]:[0] [P3]:[1]\n")
            f.write(f"\t[OUTPUT]:[TRUE]\n")
                
        f.write("[LAYER]:[TOP_OF_ROCK]\n")
        f.write("\t[OUTPUT]:[TRUE]\n")
        f.write(f"[HALFSPACE]:[ELASTIC] [UNIT_WEIGHT]:[{rock['unit_weight']}] [SHEAR]:[{rock['Vs']}] [DAMPING]:[{rock['damping']}]\n")
        f.write("[RS_TYPE]:[FREQUENCY] [RS_DAMPING]:[0.05]\n")
        f.write("[ACCELERATION_HISTORY]:[EXTERNAL] [DEEPSOILACCELINPUT.TXT]\n")
        f.write(f"[UNITS]:[{units.upper()}]\n")
        f.write(f"[LAYER_NAMES]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f"\t[LAYER_NAMES]:[{layer_ii+1}][{name_list[layer_ii].replace(' ','_')}]\n")
        f.write(f"[LAYER_VELOCITY]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f'[LAYER_VEL]:[{layer_ii+1}] [VARYING]:[FALSE] [VELOCITY_DATA_COUNTS]:[1]\n')
            f.write(f'    [VELOCITY]:[{Vs_list[layer_ii]}]\n')
        f.write(f"[LAYER_STRENGTH]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f'[LAYER_STRENGTH]:[{layer_ii+1}] [VARYING]:[FALSE] [STRENGTH_DATA_COUNTS]:[1]\n')
            f.write(f'    [STRENGTH]:[{Vs_list[layer_ii]}]\n')
    
    doc = et.Element('doc')
    info = et.SubElement(doc, 'info')
    et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
    et.SubElement(info, 'version').text = "7.0.26.0"
    ref_curves = et.SubElement(doc, 'reference_curves')

    eff_stress = 0
    for layer_ii, layer in enumerate(name_list):
        xml_layer = et.SubElement(ref_curves, f'layer{layer_ii+1}')
        eff_stress += (gamma_list[layer_ii] - 9.81) * thickness_list[layer_ii] / 2.0
        gmax = gamma_list[layer_ii] / 9.81 * Vs_list[layer_ii] ** 2.0
        et.SubElement(xml_layer, 'class').text = 'sand'
        et.SubElement(xml_layer, 'model').text = 'Darendeli'
        et.SubElement(xml_layer, 'EffectiveStress').text = str(eff_stress)
        et.SubElement(xml_layer, 'Modulus').text = str(gmax)
        et.SubElement(xml_layer, 'OCR').text = str(ocr_list[layer_ii])
        et.SubElement(xml_layer, 'K0').text = str(Ko_list[layer_ii])
        et.SubElement(xml_layer, 'N').text = str(10)
        et.SubElement(xml_layer, 'Frequency').text = str(1)
        et.SubElement(xml_layer, 'PI').text = str(pi_list[layer_ii])
        eff_stress += (gamma_list[layer_ii] - 9.81) * thickness_list[layer_ii] / 2.0
    
    mydata = et.ElementTree(doc)
    indent(doc)
    with open(os.path.join(out_dir, f'meta_Profile1'), 'wb') as f:
        mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write all meta profiles
    for analysis_ii, this_analysis in enumerate(parameters_list):
        doc = et.Element('doc')
        info = et.SubElement(doc, 'info')
        et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
        et.SubElement(info, 'version').text = "7.0.26.0"
        ref_curves = et.SubElement(doc, 'reference_curves')

        for layer_ii, layer in enumerate(elements):
            cur_main_layer = np.argmin(np.array(layers_bottom) < layer['bottom']) + 1

            xml_layer = et.SubElement(ref_curves, f'layer{layer_ii+1}')
            et.SubElement(xml_layer, 'class').text = 'sand'
            et.SubElement(xml_layer, 'model').text = 'Darendeli'
            et.SubElement(xml_layer, 'EffectiveStress').text = str(layer['Eff_Stress'])
            et.SubElement(xml_layer, 'Modulus').text = str(layer['Gmax'])
            et.SubElement(xml_layer, 'OCR').text = str(ocr_list[cur_main_layer - 1])
            et.SubElement(xml_layer, 'K0').text = str(Ko_list[cur_main_layer - 1])
            et.SubElement(xml_layer, 'N').text = str(10)
            et.SubElement(xml_layer, 'Frequency').text = str(1)
            et.SubElement(xml_layer, 'PI').text = str(pi_list[cur_main_layer - 1])
        
        mydata = et.ElementTree(doc)
        indent(doc)
        with open(os.path.join(out_dir, f'meta_Profile{analysis_ii+2}'), 'wb') as f:
            mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write profiles
    for analysis_ii, this_analysis in enumerate(parameters_list):
        with open(os.path.join(out_dir, f'Profile{analysis_ii+2}'), 'w') as f:
            f.write("[FILE_VERSION]:[1]\n")
            f.write("[ANALYSIS_DOMAIN]:[TIME+FREQUENCY]\n")
            f.write("[MAX_ITERATIONS]:[15] [COMPLEX_MOD]:[SHAKE_FI] [EFFECTIVE_SSR]:[0.65]\n")
            f.write("[ANALYSIS_TYPE]:[NONLINEAR]\n")
            f.write("[SHEAR_TYPE]:[VELOCITY]\n")
            f.write("[MAX_ITERATIONS]:[5]\n")
            f.write("[ERROR_TOL]:[1E-05]\n")
            f.write("[STEP_CONTROL]:[FLEXIBLE] [MAX_STRAIN_INC]:[0] [INTERPOLATION]:[LINEAR]\n")
            f.write("[VISCOUS_DAMPING]:[FREQ_IND]\n")
            f.write("[DAMPING_UPDATE]:[FALSE]\n")
            f.write(f"[NUM_LAYERS]:[{len(elements)}]\n")
            f.write(f"[WATER_TABLE]:[{int(water_table)}]\n")

            for layer_ii, layer in enumerate(elements):
                cur_main_layer = np.argmin(np.array(layers_bottom) < layer['bottom']) + 1
                if str(cur_main_layer) in this_analysis:
                    Vs1 = Vs1_list[cur_main_layer - 1]
                    n = n_list[cur_main_layer - 1]
                    if 'Vs1' in this_analysis[str(cur_main_layer)]:
                        Vs1 = this_analysis[str(cur_main_layer)]['Vs1']
                    if 'n' in this_analysis[str(cur_main_layer)]:
                        n = this_analysis[str(cur_main_layer)]['n']
                    layer['Vs'] = Vs1 * ((1.0+2.0*Ko_list[cur_main_layer - 1])/3.0 * layer['Eff_Stress']/101.3) ** n
                f.write(f"[LAYER]:[{layer_ii+1}]\n")
                f.write(f"\t[THICKNESS]:[{layer['thickness']}] [WEIGHT]:[{layer['unit_weight']}] [SHEAR]:[{layer['Vs']}] [SS_DAMP]:[{layer['D']}]\n")
                f.write(f"\t[MODEL]:[GQ] [STRENGTH]:[{layer['strength']}] [THETA1]:[{layer['t1']:.5f}] [THETA2]:[{layer['t2']:.5f}] [THETA3]:[{layer['t3']:.5f}] [THETA4]:[{layer['t4']:.5f}] [THETA5]:[{layer['t5']:.5f}] [A]:[1]\n")
                f.write(f"\t[MRDF]:[UIUC] [P1]:[{layer['p1']:.5f}] [P2]:[{layer['p2']:.5f}] [P3]:[{layer['p3']:.5f}]\n")
                f.write(f"\t[OUTPUT]:[TRUE]\n")
                
            f.write("[LAYER]:[TOP_OF_ROCK]\n")
            f.write("\t[OUTPUT]:[TRUE]\n")
            f.write(f"[HALFSPACE]:[ELASTIC] [UNIT_WEIGHT]:[{rock['unit_weight']}] [SHEAR]:[{rock['Vs']}] [DAMPING]:[{rock['damping']}]\n")
            f.write("[RS_TYPE]:[FREQUENCY] [RS_DAMPING]:[0.05]\n")
            f.write("[ACCELERATION_HISTORY]:[EXTERNAL] [DEEPSOILACCELINPUT.TXT]\n")
            f.write(f"[UNITS]:[{units.upper()}]\n")
            f.write(f"[LAYER_NAMES]:[{len(elements)}]\n")

            for layer_ii, layer in enumerate(elements):
                f.write(f"\t[LAYER_NAMES]:[{layer_ii+1}][{layer['Name'].replace(' ','_')}]\n")
    
    with ZipFile(os.path.join(out_dir, filename[:-5] + '.dpz'), 'w') as dpz_file:
        dpz_file.write(os.path.join(out_dir, f'Profile1'), arcname=f'Profile1')
        dpz_file.write(os.path.join(out_dir, f'meta_Profile1'), arcname=f'meta_Profile1')
        dpz_file.write(os.path.join(out_dir, f'RandomizationSettings'), arcname=f'RandomizationSettings')
        for ii, _ in enumerate(parameters_list):
            dpz_file.write(os.path.join(out_dir, f'Profile{ii+2}'), arcname=f'Profile{ii+2}')
            dpz_file.write(os.path.join(out_dir, f'meta_Profile{ii+2}'), arcname=f'meta_Profile{ii+2}')

    
    os.remove(os.path.join(out_dir, f'Profile1'))
    os.remove(os.path.join(out_dir, f'meta_Profile1'))
    os.remove(os.path.join(out_dir, f'RandomizationSettings'))
    for ii, _ in enumerate(parameters_list):
        os.remove(os.path.join(out_dir, f'Profile{ii+2}'))
        os.remove(os.path.join(out_dir, f'meta_Profile{ii+2}'))



def generate_random_velocities(filename, num_cases=40, randomize_layers=[1], calibrate_layers=[], out_dir='random_study'):
    
    ensure_dir(out_dir)
    wb = xl.load_workbook(filename, data_only=True)

    # read idealized layering boundaries
    sh = wb['Deepsoil-Input']
    layers_bottom = []
    Vs1_list = []
    n_list = []
    Ko_list = []
    ocr_list = []
    pi_list = []
    gamma_list = []
    thickness_list = []
    name_list = []
    Vs_list = []
    ref_curve_list = []
    layer_bottom = 0.0
    row_num = 2
    while not sh[f'A{row_num}'].value is None:
       thickness = float(sh[f'C{row_num}'].value)
       layer_bottom += thickness
       layers_bottom.append(layer_bottom)
       Vs1_list.append(sh[f'M{row_num}'].value)
       n_list.append(sh[f'P{row_num}'].value)
       Ko_list.append(sh[f'H{row_num}'].value)
       ocr_list.append(sh[f'I{row_num}'].value)
       pi_list.append(sh[f'J{row_num}'].value)
       gamma_list.append(sh[f'E{row_num}'].value)
       thickness_list.append(sh[f'C{row_num}'].value)
       name_list.append(sh[f'B{row_num}'].value)
       Vs_list.append(sh[f'M{row_num}'].value)
       ref_curve_list.append(str(sh[f'Q{row_num}'].value))
       row_num += 1

    sh = wb['Info']
    rock = {}
    rock['unit_weight'] = float(sh['D10'].value)
    rock['Vs'] = float(sh['D9'].value)
    rock['damping'] = float(sh['D11'].value)
    units = str(sh['D16'].value)
    
    # read current calibration
    sh = wb['Deepsoil-Calibrated']
    elements = []
    elem_bottom = 0.0
    row_num = 2
    water_table = -1
    randomize_elements = []
    while not sh[f'A{row_num}'].value is None:
       cur_elem = {}
       cur_elem['Name'] = str(sh[f'B{row_num}'].value)
       cur_elem['thickness'] = float(sh[f'C{row_num}'].value)
       elem_bottom += cur_elem['thickness']
       cur_elem['bottom'] = elem_bottom
       cur_elem['water'] = str(sh[f'D{row_num}'].value)
       if water_table == -1 and cur_elem['water'] == 'Y':
           water_table = int(sh[f'A{row_num}'].value)
       cur_elem['unit_weight'] = float(sh[f'E{row_num}'].value)
       cur_elem['Vs'] = float(sh[f'F{row_num}'].value)
       cur_elem['strength'] = float(sh[f'G{row_num}'].value)
       cur_elem['D'] = float(sh[f'H{row_num}'].value)
       cur_elem['t1'] = float(sh[f'I{row_num}'].value)
       cur_elem['t2'] = float(sh[f'J{row_num}'].value)
       cur_elem['t3'] = float(sh[f'K{row_num}'].value)
       cur_elem['t4'] = float(sh[f'L{row_num}'].value)
       cur_elem['t5'] = float(sh[f'M{row_num}'].value)
       cur_elem['p1'] = float(sh[f'N{row_num}'].value)
       cur_elem['p2'] = float(sh[f'O{row_num}'].value)
       cur_elem['p3'] = float(sh[f'P{row_num}'].value)
       cur_elem['Tot_Stress'] = float(sh[f'Q{row_num}'].value)
       cur_elem['PWP'] = float(sh[f'R{row_num}'].value)
       cur_elem['Eff_Stress'] = float(sh[f'S{row_num}'].value)
       cur_elem['Gmax'] = float(sh[f'T{row_num}'].value)
       cur_main_layer = np.argmin(np.array(layers_bottom) < cur_elem['bottom']) + 1
       if cur_main_layer in randomize_layers:
           randomize_elements.append(row_num-2)

       elements.append(cur_elem)

       row_num += 1

    # generate random variables
    norm_dist = np.zeros((len(randomize_elements), num_cases))
    for ii, _ in enumerate(randomize_elements):
        norm_dist[ii,:] = np.random.normal(0.0, 1.0, (1, num_cases))
    Vs_to_write = np.zeros((len(randomize_elements), num_cases+1))
    Vs_to_write[:,0] = np.array(randomize_elements)

    # write files to be zipped

    # wtite the 'RandomizationSettings' file
    doc = et.Element('doc')
    info = et.SubElement(doc, 'info')
    et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
    et.SubElement(info, 'version').text = "7.0.26.0"
    values = et.SubElement(doc, 'values')
    et.SubElement(values, 'b').text = str(0)
    et.SubElement(values, 'C1').text = str(10.86)
    et.SubElement(values, 'C2').text = str(0.89)
    et.SubElement(values, 'C3').text = str(1.98)
    et.SubElement(values, 'Delta').text = str(13.1)
    et.SubElement(values, 'h0').text = str(0.095)
    et.SubElement(values, 'isThicknessRandomized').text = str(True)
    et.SubElement(values, 'isVelocityRandomized').text = str(True)
    et.SubElement(values, 'isDynamicCurveRandomized').text = str(False)
    et.SubElement(values, 'max_frequency').text = str(25)
    et.SubElement(values, 'max_thickness').text = str(3)
    et.SubElement(values, 'min_thickness').text = str(0.5)
    et.SubElement(values, 'NumberOfRealizations').text = str(num_cases)
    et.SubElement(values, 'Rho0').text = str(0.96)
    et.SubElement(values, 'Rho200').text = str(0.96)
    et.SubElement(values, 'sigma').text = str(0.46)
    et.SubElement(values, 'rho1').text = str(-1)
    et.SubElement(values, 'rho2').text = str(0)
    et.SubElement(values, 'STDEVbound').text = str(1.5)
    et.SubElement(values, 'IsWaterTableChecked').text = str(True)
    et.SubElement(values, 'VRandomMethod').text = str("Geomatrix_AB")
    mydata = et.ElementTree(doc)
    indent(doc)
    with open(os.path.join(out_dir, 'RandomizationSettings'), 'wb') as f:
        mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write profile and meta profile 1
    with open(os.path.join(out_dir, f'Profile1'), 'w') as f:
        f.write("[FILE_VERSION]:[1]\n")
        f.write("[ANALYSIS_DOMAIN]:[TIME+FREQUENCY]\n")
        f.write("[MAX_ITERATIONS]:[15] [COMPLEX_MOD]:[SHAKE_FI] [EFFECTIVE_SSR]:[0.65]\n")
        f.write("[ANALYSIS_TYPE]:[NONLINEAR]\n")
        f.write("[SHEAR_TYPE]:[VELOCITY]\n")
        f.write("[MAX_ITERATIONS]:[5]\n")
        f.write("[ERROR_TOL]:[1E-05]\n")
        f.write("[STEP_CONTROL]:[FLEXIBLE] [MAX_STRAIN_INC]:[0] [INTERPOLATION]:[LINEAR]\n")
        f.write("[VISCOUS_DAMPING]:[FREQ_IND]\n")
        f.write("[DAMPING_UPDATE]:[FALSE]\n")
        f.write(f"[NUM_LAYERS]:[{len(name_list)}]\n")
        f.write("[WATER_TABLE]:[1]\n")

        for layer_ii, layer in enumerate(name_list):
            f.write(f"[LAYER]:[{layer_ii+1}]\n")
            f.write(f"\t[THICKNESS]:[{thickness_list[layer_ii]}] [WEIGHT]:[{gamma_list[layer_ii]}] [SHEAR]:[{Vs_list[layer_ii]}] [SS_DAMP]:[0.02]\n")
            f.write(f"\t[MODEL]:[GQ] [STRENGTH]:[5] [THETA1]:[-20] [THETA2]:[-20] [THETA3]:[0] [THETA4]:[1] [THETA5]:[0.1] [A]:[0]\n")
            f.write(f"\t[MRDF]:[UIUC] [P1]:[1] [P2]:[0] [P3]:[1]\n")
            f.write(f"\t[OUTPUT]:[TRUE]\n")
                
        f.write("[LAYER]:[TOP_OF_ROCK]\n")
        f.write("\t[OUTPUT]:[TRUE]\n")
        f.write(f"[HALFSPACE]:[ELASTIC] [UNIT_WEIGHT]:[{rock['unit_weight']}] [SHEAR]:[{rock['Vs']}] [DAMPING]:[{rock['damping']}]\n")
        f.write("[RS_TYPE]:[FREQUENCY] [RS_DAMPING]:[0.05]\n")
        f.write("[ACCELERATION_HISTORY]:[EXTERNAL] [DEEPSOILACCELINPUT.TXT]\n")
        f.write(f"[UNITS]:[{units.upper()}]\n")
        f.write(f"[LAYER_NAMES]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f"\t[LAYER_NAMES]:[{layer_ii+1}][{name_list[layer_ii].replace(' ','_')}]\n")
        f.write(f"[LAYER_VELOCITY]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f'[LAYER_VEL]:[{layer_ii+1}] [VARYING]:[FALSE] [VELOCITY_DATA_COUNTS]:[1]\n')
            f.write(f'    [VELOCITY]:[{Vs_list[layer_ii]}]\n')
        f.write(f"[LAYER_STRENGTH]:[{len(name_list)}]\n")
        for layer_ii, layer in enumerate(name_list):
            f.write(f'[LAYER_STRENGTH]:[{layer_ii+1}] [VARYING]:[FALSE] [STRENGTH_DATA_COUNTS]:[1]\n')
            f.write(f'    [STRENGTH]:[{Vs_list[layer_ii]}]\n')
    
    doc = et.Element('doc')
    info = et.SubElement(doc, 'info')
    et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
    et.SubElement(info, 'version').text = "7.0.26.0"
    ref_curves = et.SubElement(doc, 'reference_curves')

    eff_stress = 0
    for layer_ii, layer in enumerate(name_list):
        xml_layer = et.SubElement(ref_curves, f'layer{layer_ii+1}')
        eff_stress += (gamma_list[layer_ii] - 9.81) * thickness_list[layer_ii] / 2.0
        gmax = gamma_list[layer_ii] / 9.81 * Vs_list[layer_ii] ** 2.0
        et.SubElement(xml_layer, 'class').text = 'sand'
        et.SubElement(xml_layer, 'model').text = 'Darendeli'
        et.SubElement(xml_layer, 'EffectiveStress').text = str(eff_stress)
        et.SubElement(xml_layer, 'Modulus').text = str(gmax)
        et.SubElement(xml_layer, 'OCR').text = str(ocr_list[layer_ii])
        et.SubElement(xml_layer, 'K0').text = str(Ko_list[layer_ii])
        et.SubElement(xml_layer, 'N').text = str(10)
        et.SubElement(xml_layer, 'Frequency').text = str(1)
        et.SubElement(xml_layer, 'PI').text = str(pi_list[layer_ii])
        eff_stress += (gamma_list[layer_ii] - 9.81) * thickness_list[layer_ii] / 2.0
    
    mydata = et.ElementTree(doc)
    indent(doc)
    with open(os.path.join(out_dir, f'meta_Profile1'), 'wb') as f:
        mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write all meta profiles
    for analysis_ii in range(num_cases):
        doc = et.Element('doc')
        info = et.SubElement(doc, 'info')
        et.SubElement(info, 'assembly').text = "DEEPSOIL 7.0"
        et.SubElement(info, 'version').text = "7.0.26.0"
        ref_curves = et.SubElement(doc, 'reference_curves')

        for layer_ii, layer in enumerate(elements):
            cur_main_layer = np.argmin(np.array(layers_bottom) < layer['bottom']) + 1

            xml_layer = et.SubElement(ref_curves, f'layer{layer_ii+1}')
            et.SubElement(xml_layer, 'class').text = 'sand'
            et.SubElement(xml_layer, 'model').text = 'Darendeli'
            et.SubElement(xml_layer, 'EffectiveStress').text = str(layer['Eff_Stress'])
            et.SubElement(xml_layer, 'Modulus').text = str(layer['Gmax'])
            et.SubElement(xml_layer, 'OCR').text = str(ocr_list[cur_main_layer - 1])
            et.SubElement(xml_layer, 'K0').text = str(Ko_list[cur_main_layer - 1])
            et.SubElement(xml_layer, 'N').text = str(10)
            et.SubElement(xml_layer, 'Frequency').text = str(1)
            et.SubElement(xml_layer, 'PI').text = str(pi_list[cur_main_layer - 1])
        
        mydata = et.ElementTree(doc)
        indent(doc)
        with open(os.path.join(out_dir, f'meta_Profile{analysis_ii+2}'), 'wb') as f:
            mydata.write(f, encoding='utf-8', xml_declaration=True)

    # write profiles
    for analysis_ii in range(num_cases):
        with open(os.path.join(out_dir, f'Profile{analysis_ii+2}'), 'w') as f:
            f.write("[FILE_VERSION]:[1]\n")
            f.write("[ANALYSIS_DOMAIN]:[TIME+FREQUENCY]\n")
            f.write("[MAX_ITERATIONS]:[15] [COMPLEX_MOD]:[SHAKE_FI] [EFFECTIVE_SSR]:[0.65]\n")
            f.write("[ANALYSIS_TYPE]:[NONLINEAR]\n")
            f.write("[SHEAR_TYPE]:[VELOCITY]\n")
            f.write("[MAX_ITERATIONS]:[5]\n")
            f.write("[ERROR_TOL]:[1E-05]\n")
            f.write("[STEP_CONTROL]:[FLEXIBLE] [MAX_STRAIN_INC]:[0] [INTERPOLATION]:[LINEAR]\n")
            f.write("[VISCOUS_DAMPING]:[FREQ_IND]\n")
            f.write("[DAMPING_UPDATE]:[FALSE]\n")
            f.write(f"[NUM_LAYERS]:[{len(elements)}]\n")
            f.write(f"[WATER_TABLE]:[{int(water_table)}]\n")

            for layer_ii, layer in enumerate(elements):
                cur_main_layer = np.argmin(np.array(layers_bottom) < layer['bottom']) + 1 
                t1,t2,t3,t4,t5,p1,p2,p3,D = layer['t1'],layer['t2'],layer['t3'],layer['t4'],layer['t5'],layer['p1'],layer['p2'],layer['p3'],layer['D']
                Vs, Gmax = layer['Vs'], layer['Gmax']
                if layer_ii in randomize_elements:
                    Vs = np.exp(np.log(layer['Vs']) + 0.10*norm_dist[layer_ii, analysis_ii])
                    Gmax = layer['unit_weight'] / 9.81 * Vs ** 2.0
                    Vs_to_write[layer_ii, analysis_ii+1] = Vs
                if cur_main_layer in calibrate_layers:
                    p = (1+2.0*Ko_list[cur_main_layer-1])/3.0 * layer['Eff_Stress']
                    initial_guess = (layer['t1'],layer['t2'],layer['t3'],layer['t4'],layer['t5'],layer['p1'],layer['p2'],layer['p3'])
                    if ref_curve_list[cur_main_layer-1] == 'Darendeli':
                        t1,t2,t3,t4,t5,p1,p2,p3,D = \
                            calibrateGQH_Darendeli_LS(p, pi_list[cur_main_layer-1], ocr_list[cur_main_layer-1], layer['strength'], Gmax, Patm=101.3, initial_guess=initial_guess)
                    elif ref_curve_list[cur_main_layer-1] == 'Zhang2005':
                         t1,t2,t3,t4,t5,p1,p2,p3,D = \
                            calibrateGQH_Zhang2005_LS(p, pi_list[cur_main_layer-1], layer['strength'], Gmax, Patm=101.3, initial_guess=initial_guess)
                    # elif ref_curve_list[cur_main_layer-1] == 'Menq':
                    #     t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_Menq_LS(p, xl_layer['Cu'], xl_layer['D50'], strength, Gmax, Patm=101.3)
                    # elif ref_curve_list[cur_main_layer-1] == 'VuceticDobry':
                    #     t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_VuceticDobry_LS(xl_layer['PI'], strength, Gmax)
                    else:
                        raise('Modulus Reduction curve ' +  ref_curve_list[cur_main_layer-1] + ' not implemented.')
                f.write(f"[LAYER]:[{layer_ii+1}]\n")
                f.write(f"\t[THICKNESS]:[{layer['thickness']}] [WEIGHT]:[{layer['unit_weight']}] [SHEAR]:[{Vs}] [SS_DAMP]:[{D}]\n")
                f.write(f"\t[MODEL]:[GQ] [STRENGTH]:[{layer['strength']}] [THETA1]:[{t1:.5f}] [THETA2]:[{t2:.5f}] [THETA3]:[{t3:.5f}] [THETA4]:[{t4:.5f}] [THETA5]:[{t5:.5f}] [A]:[1]\n")
                f.write(f"\t[MRDF]:[UIUC] [P1]:[{p1:.5f}] [P2]:[{p2:.5f}] [P3]:[{p3:.5f}]\n")
                f.write(f"\t[OUTPUT]:[TRUE]\n")
                
            f.write("[LAYER]:[TOP_OF_ROCK]\n")
            f.write("\t[OUTPUT]:[TRUE]\n")
            f.write(f"[HALFSPACE]:[ELASTIC] [UNIT_WEIGHT]:[{rock['unit_weight']}] [SHEAR]:[{rock['Vs']}] [DAMPING]:[{rock['damping']}]\n")
            f.write("[RS_TYPE]:[FREQUENCY] [RS_DAMPING]:[0.05]\n")
            f.write("[ACCELERATION_HISTORY]:[EXTERNAL] [DEEPSOILACCELINPUT.TXT]\n")
            f.write(f"[UNITS]:[{units.upper()}]\n")
            f.write(f"[LAYER_NAMES]:[{len(elements)}]\n")

            for layer_ii, layer in enumerate(elements):
                f.write(f"\t[LAYER_NAMES]:[{layer_ii+1}][{layer['Name'].replace(' ','_')}]\n")
    np.savetxt(os.path.join(out_dir, 'Generated_Vs.csv'), Vs_to_write, delimiter=',')

    
    with ZipFile(os.path.join(out_dir, filename[:-5] + '.dpz'), 'w') as dpz_file:
        dpz_file.write(os.path.join(out_dir, f'Profile1'), arcname=f'Profile1')
        dpz_file.write(os.path.join(out_dir, f'meta_Profile1'), arcname=f'meta_Profile1')
        dpz_file.write(os.path.join(out_dir, f'RandomizationSettings'), arcname=f'RandomizationSettings')
        for ii in range(num_cases):
            dpz_file.write(os.path.join(out_dir, f'Profile{ii+2}'), arcname=f'Profile{ii+2}')
            dpz_file.write(os.path.join(out_dir, f'meta_Profile{ii+2}'), arcname=f'meta_Profile{ii+2}')

    
    os.remove(os.path.join(out_dir, f'Profile1'))
    os.remove(os.path.join(out_dir, f'meta_Profile1'))
    os.remove(os.path.join(out_dir, f'RandomizationSettings'))
    for ii in range(num_cases):
        os.remove(os.path.join(out_dir, f'Profile{ii+2}'))
        os.remove(os.path.join(out_dir, f'meta_Profile{ii+2}'))




def main_parameter_study():
    # generate_parameter_study('Bhr-01-Simplified_new.xlsx', out_dir='Bhr-01_parameter_study_temp')
    # generate_random_velocities('Bhr-01-Simplified_new.xlsx', 40, calibrate_layers=[], out_dir='Bhr-01_parameter_study_temp')
    generate_random_velocities('Bhr-01-Simplified_new.xlsx', 40, calibrate_layers=[], out_dir='Bhr-01_parameter_study_temp2')


if __name__ == "__main__":
    main_parameter_study()