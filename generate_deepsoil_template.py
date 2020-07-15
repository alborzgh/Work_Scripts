import os
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl

import sys
sys.path.insert(1, './Utility_Scripts')

from os_utils import ensure_dir
from calibrateGQH import calibrateGQH_Darendeli_LS, calibrateGQH_Menq_LS, calibrateGQH_VuceticDobry_LS,calibrateGQH_Zhang2005_LS
from ModulusReduction import Menq, Darendeli, VuceticDobry, Zhang2005
from GQHModel import GQH_modulusreduction, GQH_damping
from subprocess import Popen, PIPE
from multiprocessing import  Process, Pool

DS_RUNTIME = r"C:\Program Files\University of Illinois at Urbana-Champaign\DEEPSOIL 7.0\soil64.exe"

def plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D, strength, Gmax, model, p=101.3, PI_Cu=0.0, OCR_D50=1.0, filename='test.png'):
    ax_title = ""
    if model == 'VuceticDobry':
        GGmax, damp, gamma = VuceticDobry(PI_Cu)
        ax_title = f"Vucetic-Dobry: PI = {PI_Cu}"
    elif model == 'Darendeli':
        GGmax, damp, gamma = Darendeli(p, PI_Cu, OCR_D50, Patm=101.3)
        ax_title = f"Darendeli: p = {p:.2f}kPa, PI = {PI_Cu}, OCR = {OCR_D50}"
    elif model == 'Menq':
        GGmax, damp, gamma = Menq(p, PI_Cu, OCR_D50, Patm=101.3)
        ax_title = f"Menq: p = {p:.2f}kPa, Cu = {PI_Cu}, D50 = {OCR_D50}"
    elif model == 'Zhang':
        GGmax, damp, gamma = Zhang2005(p, PI_Cu, age='Quaternary', Patm=101.3)

        ax_title = f"Zhang2005: p = {p:.2f}kPa, Cu = {PI_Cu}, age='Quaternary'"
    else:
        raise(f"Material model {model} is not implemented!")

    GQH_GGmax = GQH_modulusreduction(strength,Gmax, t1, t2, t3, t4, t5, gamma)
    GQH_damp_org  = GQH_damping(strength, Gmax, t1, t2, t3, t4, t5, gamma)
    factor = p1 - p2 * (1.0 - GQH_GGmax)**p3
    GQH_damp = D + factor * GQH_damp_org

    fig = plt.figure(figsize=(7,8), dpi=150, tight_layout=True)
    ax1 = fig.add_subplot(2,1,1)
    ax2 = fig.add_subplot(2,1,2)
    ax3 = ax1.twinx()
    ax1.semilogx(gamma*100.0, GGmax, 'b-', lw=1.0, label='Empirical')
    ax1.semilogx(gamma*100.0, GQH_GGmax, 'r-', lw=2.0, label='GQH')
    ax3.semilogx(gamma*100.0, damp, 'b--', lw=1.0)
    ax3.semilogx(gamma*100.0, GQH_damp*100.0, 'r--', lw=2.0)
    ax1.set_xlim(0.0001,10.0)
    ax1.set_ylim(0.0,1.0)
    ax3.set_ylim(ymin=0.0)
    ax1.set_xlabel(r'$\gamma$ (%)')
    ax1.set_ylabel(r'$G/G_{max}$')
    ax3.set_ylabel(r'$\xi$ (%)')
    ax1.set_title(ax_title)
    ax1.grid(which='both', lw=0.25, linestyle=':', color=(0.25,0.25,0.25))
    
    ax2.plot(gamma*100.0, GGmax*Gmax*gamma, 'b-', lw=1.0, label='Empirical')
    ax2.plot(gamma*100.0, GQH_GGmax*Gmax*gamma, 'r-', lw=2.0, label='GQH')
    ax2.axhline(strength, color='k', linestyle='--', lw=1.0)
    ax2.set_xlabel(r'$\gamma$ (%)')
    ax2.set_ylabel(r'$\tau$ (kPa)')
    ax2.set_xlim(0.0,10.0)
    ax2.set_ylim(ymin=0.0)
    ax2.legend(loc="lower right")
    ax2.grid(which='both', lw=0.25, linestyle=':', color=(0.25,0.25,0.25))

    fig.savefig(filename)
    plt.close(fig)

def read_excel_file(filename):
    wb = xl.load_workbook(filename, data_only=True)
    sh = wb['Deepsoil-Input']

    layers = []

    # find number of layers
    numLayers = 0
    while True:
        if sh[f"A{numLayers+2}"].value == None:
            break
        else:
            cur_layer= {}
            cur_layer['Name'] = str(sh[f"B{numLayers+2}"].value)
            cur_layer['thickness'] = float(sh[f"C{numLayers+2}"].value)
            cur_layer['water'] = str(sh[f"D{numLayers+2}"].value)
            cur_layer['unit_weight'] = float(sh[f"E{numLayers+2}"].value)
            cur_layer['friction_angle'] = float(sh[f"F{numLayers+2}"].value)
            cur_layer['cohesion'] = float(sh[f"G{numLayers+2}"].value)
            cur_layer['K0'] = float(sh[f"H{numLayers+2}"].value)
            cur_layer['OCR'] = float(sh[f"I{numLayers+2}"].value)
            cur_layer['PI'] = float(sh[f"J{numLayers+2}"].value)
            cur_layer['D50'] = float(sh[f"K{numLayers+2}"].value)
            cur_layer['Cu'] = float(sh[f"L{numLayers+2}"].value)
            cur_layer['Vs1'] = float(sh[f"M{numLayers+2}"].value)
            cur_layer['Vs2'] = float(sh[f"N{numLayers+2}"].value)
            cur_layer['Vs_interp'] = str(sh[f"O{numLayers+2}"].value)
            cur_layer['Vs3'] = float(sh[f"P{numLayers+2}"].value)
            cur_layer['MR_curve'] = str(sh[f"Q{numLayers+2}"].value)
            cur_layer['damping'] = float(sh[f"R{numLayers+2}"].value)
            cur_layer['max_frequency'] = float(sh[f"S{numLayers+2}"].value)
            cur_layer['descretize'] = str(sh[f"T{numLayers+2}"].value)
            
            cur_layer['finecontent'] = float(sh[f"U{numLayers+2}"].value)
            
            layers.append(cur_layer)
            numLayers += 1
    
    gwt = float(wb['Info']['D7'].value)

    rock = {}
    rock['Vs'] = float(wb['Info']['D9'].value)
    rock['unit_weight'] = float(wb['Info']['D10'].value)
    rock['damping'] = float(wb['Info']['D11'].value)

    units = str(wb['Info']['D16'].value)

    return layers, gwt, rock, units

def generate_deepsoil_layering(xl_fn, filename):

    xl_layers, gwt, rock, units = read_excel_file(xl_fn)

    ds_layers = []
    depth = 0.0
    tot_stress = 0.0
    pwp = 0.0
    eff_stress = 0.0
    water_index = 0
    
    plots_dir = filename[:-3] + '_plots' 
    ensure_dir(plots_dir)

    for xl_layer_ii, xl_layer in enumerate(xl_layers):
        if xl_layer['descretize'] == 'Y':
            Vs_min = xl_layer['Vs1']
            num_sublayers = int(np.ceil(xl_layer['thickness'] / (Vs_min / 4/xl_layer['max_frequency'])))
            if xl_layer['Vs_interp'] == 'Linear' and xl_layer['Vs2'] < xl_layer['Vs1']:
                Vs_min = xl_layer['Vs2']
                num_sublayers = int(np.ceil(xl_layer['thickness'] / (Vs_min / 4*xl_layer['max_frequency'])))
            elif xl_layer['Vs_interp'] == 'Stress-Dependent':
                num_sublayers = 1
                while True:
                    tot_Vs_stress = tot_stress + xl_layer['thickness'] / num_sublayers * xl_layer['unit_weight']
                    pwp_Vs_stress = 0.0
                    if depth + xl_layer['thickness'] / num_sublayers >= gwt:
                        pwp_Vs_stress = 9.81 * (depth + xl_layer['thickness'] / num_sublayers / 2.0 - gwt)
                    eff_Vs_stress = tot_Vs_stress - pwp_Vs_stress
                    p_Vs = (1.0 + 2.0 * xl_layer['K0']) / 3.0 * eff_Vs_stress
                    Vs_test = xl_layer['Vs1'] * (p_Vs/101.3)**xl_layer['Vs3']
                    if Vs_test / (4 * xl_layer['thickness'] / num_sublayers) > xl_layer['max_frequency']:
                        break
                    else:
                        num_sublayers *= 2
            elif xl_layer['Vs_interp'] == 'PRDP-2020':
                num_sublayers = 1
                while True:
                    depth_test = depth + xl_layer['thickness'] / num_sublayers
                    # Vs_test = 0.047*depth_test**3.0 - 1.69*depth_test**2.0 + 23.0*depth_test + 115.0
                    # Vs_test = 0.055*depth_test**3.0 - 1.77*depth_test**2.0 + 21.7*depth_test + 117.0
                    
                    # Vs_test = 0.051*depth_test**3.0 - 1.73*depth_test**2.0 + 22.3*depth_test + 116.0
                    # Vs_test = 0.045*depth_test**3.0 - 1.63*depth_test**2.0 + 22.1*depth_test + 116.0
                    
                    # Vs_test = 0.0167*depth_test**3.0 - 0.94*depth_test**2.0 + 17.57*depth_test + 122.0
                    Vs_test = -0.0006*depth_test**4.0 +0.0507*depth_test**3- 1.5814*depth_test**2.0 + 22.16*depth_test + 116.0
                    
                    if Vs_test / (4 * xl_layer['thickness'] / num_sublayers) > xl_layer['max_frequency']:
                        break
                    else:
                        num_sublayers *= 2
            elif xl_layer['Vs_interp'] == 'JD-2018-Fill':
                num_sublayers = 1
                while True:
                    depth_test = depth + xl_layer['thickness'] / num_sublayers
                    Vs_test = 0.018*depth_test**3.0 - 0.89*depth_test**2.0 + 16.2*depth_test + 104.0
                    if Vs_test / (4 * xl_layer['thickness'] / num_sublayers) > xl_layer['max_frequency']:
                        break
                    else:
                        num_sublayers *= 2
            elif xl_layer['Vs_interp'] == 'JD-2018-SoftClay':
                num_sublayers = 1
                while True:
                    depth_test = depth + xl_layer['thickness'] / num_sublayers
                    Vs_test = 1.5*depth + 117.5
                    if Vs_test / (4 * xl_layer['thickness'] / num_sublayers) > xl_layer['max_frequency']:
                        break
                    else:
                        num_sublayers *= 2

            sublayer_thickness = xl_layer['thickness'] / num_sublayers
            for layer_ii in range(num_sublayers):
                depth += sublayer_thickness / 2.0
                tot_stress += xl_layer['unit_weight'] * sublayer_thickness / 2.0
                if depth >= gwt:
                    pwp += 9.81 * sublayer_thickness / 2.0
                eff_stress = tot_stress - pwp
                p = (1.0 + 2.0 * xl_layer['K0']) / 3.0 * eff_stress
                strength = xl_layer['cohesion'] + eff_stress * np.tan(np.pi/180.0*xl_layer['friction_angle'])
                if xl_layer['Vs_interp'] == 'Constant':
                    Vs = xl_layer['Vs1']
                elif xl_layer['Vs_interp'] == 'Linear':
                    Vs = xl_layer['Vs1'] + layer_ii * (xl_layer['Vs2']-xl_layer['Vs1']) / num_sublayers
                elif xl_layer['Vs_interp'] == 'Stress-Dependent':
                    Vs = xl_layer['Vs1'] * (p/101.3)**xl_layer['Vs3']
                elif xl_layer['Vs_interp'] == 'JD-2018-Fill':
                    Vs = 0.018*depth**3.0 - 0.89*depth**2.0 + 16.2*depth + 104.0
                elif xl_layer['Vs_interp'] == 'PRDP-2020':
                    # Vs = 0.047*depth**3.0 - 1.69*depth**2.0 + 23*depth + 115.0
                    # Vs = 0.055*depth**3.0 - 1.77*depth**2.0 + 21.7*depth + 117.0
                    
                    # Vs = 0.051*depth**3.0 - 1.73*depth**2.0 + 22.3*depth + 116.0
                    # Vs = 0.045*depth**3.0 - 1.63*depth**2.0 + 22.1*depth + 116.0
                    # Vs = 0.0167*depth**3.0 - 0.94*depth**2.0 + 17.57*depth + 122.0
                    Vs = -0.0006*depth**4.0 +0.0507*depth**3- 1.5814*depth**2.0 + 22.16*depth + 116.0
                elif xl_layer['Vs_interp'] == 'JD-2018-SoftClay':
                    Vs = 1.5*depth + 117.5
                else:
                    raise('Vs interpolation ' + xl_layer['Vs_interp'] + ' not implemented.')
                
                Gmax = xl_layer['unit_weight'] / 9.81 * Vs**2.0

                if xl_layer['MR_curve'] == 'Darendeli':
                    t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_Darendeli_LS(p, xl_layer['PI'], xl_layer['OCR'], strength, Gmax, Patm=101.3)
                    plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Darendeli', p, xl_layer['PI'], xl_layer['OCR'], plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name'] + '_' + str(layer_ii+1)+'.png')
                elif xl_layer['MR_curve'] == 'Menq':
                    t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_Menq_LS(p, xl_layer['Cu'], xl_layer['D50'], strength, Gmax, Patm=101.3)
                    plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Menq', p, xl_layer['Cu'], xl_layer['D50'], plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name'] + '_' + str(layer_ii+1)+'.png')
                elif xl_layer['MR_curve'] == 'VuceticDobry':
                    t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_VuceticDobry_LS(xl_layer['PI'], strength, Gmax)
                    plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'VuceticDobry', p, 0.0, 0.0, plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name'] + '_' + str(layer_ii+1)+'.png')
                elif xl_layer['MR_curve'] == 'Zhang':
                    t1,t2,t3,t4,t5,p1,p2,p3,D = \
                            calibrateGQH_Zhang2005_LS(p, xl_layer['PI'], strength, Gmax, Patm=101.3)
                    
                    plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Zhang', p, 0.0, 0.0, plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name'] + '_' + str(layer_ii+1)+'.png')
                else:
                    raise('Modulus Reduction curve ' +  xl_layer['MR_curve'] + ' not implemented.')
                
                if  depth >= gwt and water_index == 0:
                    water_index = (len(ds_layers) + 1)
                
                cur_layer= {}
                cur_layer['Name'] = xl_layer['Name'] + '_' + str(layer_ii+1)
                cur_layer['thickness'] = sublayer_thickness
                cur_layer['water'] = 'Y' if depth >= gwt else 'N'
                cur_layer['unit_weight'] = xl_layer['unit_weight']
                cur_layer['Vs'] = Vs
                cur_layer['strength'] = strength
                cur_layer['D'] = D
                cur_layer['t1'] = t1
                cur_layer['t2'] = t2
                cur_layer['t3'] = t3
                cur_layer['t4'] = t4
                cur_layer['t5'] = t5
                cur_layer['p1'] = p1
                cur_layer['p2'] = p2
                cur_layer['p3'] = p3
                cur_layer['Gmax'] = Gmax
                cur_layer['Tot_Stress'] = tot_stress
                cur_layer['Eff_Stress'] = eff_stress
                cur_layer['PWP'] = pwp
                cur_layer['finecontent'] = xl_layer['finecontent'] 
                ds_layers.append(cur_layer)
                depth += sublayer_thickness / 2.0
                tot_stress += xl_layer['unit_weight'] * sublayer_thickness / 2.0
                if depth >= gwt:
                    pwp += 9.81 * sublayer_thickness / 2.0
                
        else:
            depth += xl_layer['thickness'] / 2.0
            tot_stress += xl_layer['unit_weight'] * xl_layer['thickness'] / 2.0
            if xl_layer['water'] == 'Y':
                pwp += 9.81 * xl_layer['thickness'] / 2.0
            eff_stress = tot_stress - pwp
            p = (1.0 + 2.0 * xl_layer['K0']) / 3.0 * eff_stress
            strength = xl_layer['cohesion'] + eff_stress * np.tan(np.pi/180.0*xl_layer['friction_angle'])
            if xl_layer['Vs_interp'] == 'Constant':
                Vs = xl_layer['Vs1']
            elif xl_layer['Vs_interp'] == 'Linear':
                Vs =0.5 * (xl_layer['Vs2']+xl_layer['Vs1'])
            elif xl_layer['Vs_interp'] == 'Stress-Dependent':
                Vs = xl_layer['Vs1'] * (p/101.3)**xl_layer['Vs3']
            elif xl_layer['Vs_interp'] == 'PRDP-2020':
                # Vs = 0.047*depth**3.0 - 1.69*depth**2.0 + 23*depth + 115.0
                # Vs = 0.055*depth**3.0 - 1.77*depth**2.0 + 21.7*depth + 117.0
                
                # Vs = 0.051*depth**3.0 - 1.73*depth**2.0 + 22.3*depth + 116.0
                # Vs = 0.045*depth**3.0 - 1.63*depth**2.0 + 22.1*depth + 116.0
                
                # Vs = 0.0167*depth**3.0 - 0.94*depth**2.0 + 17.57*depth + 122.0
                
                Vs = -0.0006*depth**4.0 +0.0507*depth**3- 1.5814*depth**2.0 + 22.16*depth + 116.0
                 
            elif xl_layer['Vs_interp'] == 'JD-2018-Fill':
                Vs = 0.018*depth**3.0 - 0.89*depth**2.0 + 16.2*depth + 104.0
            elif xl_layer['Vs_interp'] == 'JD-2018-SoftClay':
                Vs = 1.5*depth + 117.5
            else:
                raise('Vs interpolation ' + xl_layer['Vs_interp'] + ' not implemented.')
                
            Gmax = xl_layer['unit_weight'] / 9.81 * Vs**2.0

            if xl_layer['MR_curve'] == 'Darendeli':
                t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_Darendeli_LS(p, xl_layer['PI'], xl_layer['OCR'], strength, Gmax, Patm=101.3)
                plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Darendeli', p, xl_layer['PI'], xl_layer['OCR'], plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name']+'.png')
            elif xl_layer['MR_curve'] == 'Menq':
                t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_Menq_LS(p, xl_layer['Cu'], xl_layer['D50'], strength, Gmax, Patm=101.3)
                plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Menq', p, xl_layer['Cu'], xl_layer['D50'], plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name']+'.png')
            elif xl_layer['MR_curve'] == 'VuceticDobry':
                t1, t2, t3, t4, t5, p1, p2, p3, D = calibrateGQH_VuceticDobry_LS(xl_layer['PI'], strength, Gmax)
                plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'VuceticDobry', p, 0.0, 0.0, plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name']+'.png')
            elif xl_layer['MR_curve'] == 'Zhang':
                t1,t2,t3,t4,t5,p1,p2,p3,D = \
                            calibrateGQH_Zhang2005_LS(p, xl_layer['PI'], strength, Gmax, Patm=101.3)
                    
                plot_curve(t1,t2,t3,t4,t5,p1,p2,p3,D,strength, Gmax, 'Zhang', p, 0.0, 0.0, plots_dir + '/' + f'{xl_layer_ii+1}_' + xl_layer['Name'] + '_' + str(layer_ii+1)+'.png')
            else:
                raise('Modulus Reduction curve ' +  xl_layer['MR_curve'] + ' not implemented.')
            
            if  depth >= gwt and water_index == 0:
                water_index = (len(ds_layers) + 1)

            cur_layer= {}
            cur_layer['Name'] = xl_layer['Name']
            cur_layer['thickness'] = xl_layer['thickness']
            cur_layer['water'] = xl_layer['water']
            cur_layer['unit_weight'] = xl_layer['unit_weight']
            cur_layer['Vs'] = Vs
            cur_layer['strength'] = strength
            cur_layer['D'] = D
            cur_layer['t1'] = t1
            cur_layer['t2'] = t2
            cur_layer['t3'] = t3
            cur_layer['t4'] = t4
            cur_layer['t5'] = t5
            cur_layer['p1'] = p1
            cur_layer['p2'] = p2
            cur_layer['p3'] = p3
            cur_layer['Gmax'] = Gmax
            cur_layer['Tot_Stress'] = tot_stress
            cur_layer['Eff_Stress'] = eff_stress
            cur_layer['PWP'] = pwp
            cur_layer['finecontent'] = xl_layer['finecontent'] 
			
            ds_layers.append(cur_layer)
            depth += xl_layer['thickness'] / 2.0
            tot_stress += xl_layer['unit_weight'] * sublayer_thickness / 2.0
            if depth >= gwt:
                pwp += 9.81 * sublayer_thickness / 2.0
    write_deepsoil_excel(ds_layers, xl_fn)
    return ds_layers, water_index, rock, units

def write_deepsoil_excel(ds_layers, xl_fn):
    wb = xl.load_workbook(xl_fn)
    sh = wb.create_sheet('Deepsoil-Calibrated')
    sh['A1'].value = 'Layer No.'
    sh['B1'].value = 'Layer Name'
    sh['C1'].value = 'Thickness'
    sh['D1'].value = 'Below Water?'
    sh['E1'].value = 'Unit Weight'
    sh['F1'].value = 'Vs'
    sh['G1'].value = 'Strength'
    sh['H1'].value = 'D'
    sh['I1'].value = 'theta 1'
    sh['J1'].value = 'theta 2'
    sh['K1'].value = 'theta 3'
    sh['L1'].value = 'theta 4'
    sh['M1'].value = 'theta 5'
    sh['N1'].value = 'p1'
    sh['O1'].value = 'p2'
    sh['P1'].value = 'p3'
    sh['Q1'].value = 'Total Stress'
    sh['R1'].value = 'PWP'
    sh['S1'].value = 'Effective Stress'
    sh['T1'].value = 'Gmax'
    sh['U1'].value = 'Fine Content'
    row_num = 2
    for layer in ds_layers:
        sh[f'A{row_num}'].value = row_num - 1
        sh[f'B{row_num}'].value = layer['Name']
        sh[f'C{row_num}'].value = layer['thickness']
        sh[f'D{row_num}'].value = layer['water']
        sh[f'E{row_num}'].value = layer['unit_weight']
        sh[f'F{row_num}'].value = layer['Vs']
        sh[f'G{row_num}'].value = layer['strength']
        sh[f'H{row_num}'].value = layer['D']
        sh[f'I{row_num}'].value = layer['t1']
        sh[f'J{row_num}'].value = layer['t2']
        sh[f'K{row_num}'].value = layer['t3']
        sh[f'L{row_num}'].value = layer['t4']
        sh[f'M{row_num}'].value = layer['t5']
        sh[f'N{row_num}'].value = layer['p1']
        sh[f'O{row_num}'].value = layer['p2']
        sh[f'P{row_num}'].value = layer['p3']
        sh[f'Q{row_num}'].value = layer['Tot_Stress']
        sh[f'R{row_num}'].value = layer['PWP']
        sh[f'S{row_num}'].value = layer['Eff_Stress']
        sh[f'T{row_num}'].value = layer['Gmax']
        sh[f'U{row_num}'].value = layer['finecontent']
        row_num += 1
    wb.save(xl_fn)


def write_deepsoil(xl_fn, filename):
    ds_layers, water_index, rock, units = generate_deepsoil_layering(xl_fn, filename)

    plots_dir = filename[:-3] + '_plots' 
    ensure_dir(plots_dir)

    with open(filename, 'w') as f:
        f.write("[FILE_VERSION]:[1]\n")
        f.write("[ANALYSIS_DOMAIN]:[TIME+FREQUENCY]\n")
        f.write("[MAX_ITERATIONS]:[15] [COMPLEX_MOD]:[SHAKE_FI] [EFFECTIVE_SSR]:[0.65]\n")
        f.write("[ANALYSIS_TYPE]:[NONLINEAR]\n")
        f.write("[SHEAR_TYPE]:[VELOCITY]\n")
        f.write("[MAX_ITERATIONS]:[5]\n")
        f.write("[ERROR_TOL]:[1E-05]\n")
        f.write("[STEP_CONTROL]:[FLEXIBLE] [MAX_STRAIN_INC]:[0] [INTERPOLATION]:[LINEAR]\n")
        f.write("[VISCOUS_DAMPING]:[FREQ_IND]\n")
        f.write("[DAMPING_UPDATE]:[FALSE]\n")
        f.write(f"[NUM_LAYERS]:[{len(ds_layers)}]\n")
        f.write(f"[WATER_TABLE]:[{int(water_index)}]\n")

        for layer_ii, layer in enumerate(ds_layers):
            f.write(f"[LAYER]:[{layer_ii+1}]\n")
            f.write(f"\t[THICKNESS]:[{layer['thickness']}] [WEIGHT]:[{layer['unit_weight']}] [SHEAR]:[{layer['Vs']}] [SS_DAMP]:[{layer['D']}]\n")
            f.write(f"\t[MODEL]:[GQ] [STRENGTH]:[{layer['strength']}] [THETA1]:[{layer['t1']:.5f}] [THETA2]:[{layer['t2']:.5f}] [THETA3]:[{layer['t3']:.5f}] [THETA4]:[{layer['t4']:.5f}] [THETA5]:[{layer['t5']:.5f}] [A]:[1]\n")
            f.write(f"\t[MRDF]:[UIUC] [P1]:[{layer['p1']:.5f}] [P2]:[{layer['p2']:.5f}] [P3]:[{layer['p3']:.5f}]\n")
            f.write(f"\t[OUTPUT]:[TRUE]\n")
            
        f.write("[LAYER]:[TOP_OF_ROCK]\n")
        f.write("\t[OUTPUT]:[TRUE]\n")
        f.write(f"[HALFSPACE]:[ELASTIC] [UNIT_WEIGHT]:[{rock['unit_weight']}] [SHEAR]:[{rock['Vs']}] [DAMPING]:[{rock['damping']}]\n")
        f.write("[RS_TYPE]:[FREQUENCY] [RS_DAMPING]:[0.05]\n")
        f.write("[ACCELERATION_HISTORY]:[EXTERNAL] [DEEPSOILACCELINPUT.TXT]\n")
        f.write(f"[UNITS]:[{units.upper()}]\n")
        f.write(f"[LAYER_NAMES]:[{len(ds_layers)}]\n")

        for layer_ii, layer in enumerate(ds_layers):
            f.write(f"\t[LAYER_NAMES]:[{layer_ii+1}][{layer['Name'].replace(' ','_')}]\n")

def generate_one_profile(profile_name):
    write_deepsoil(profile_name + '.xlsx', profile_name + '.dp')
    # write_soil64(profile_name + '.dp','../Motions/Matched_Duzce_531_N.txt')

def parallel_generate():
    profiles = [
        
        "../Bhr-04/Bhr-04-Darendeli_soft1",
        # "../Bhr-04/Bhr-04-Darendeli_soft2",
        "../Bhr-04/Bhr-04-Darendeli_soft3",
        "../Bhr-04/Bhr-04-Zhang_soft1",
        # "../Bhr-04/Bhr-04-Zhang_soft2",
        "../Bhr-04/Bhr-04-Zhang_soft3",
        
        # "../Bhd-01/Bhd-01-Darendeli_soft2",
        # "../Bhd-01/Bhd-01-Zhang_soft2",
        
        # "../Bhr-01/Bhr-01-Darendeli_soft2",
        # "../Bhr-01/Bhr-01-Zhang_soft2",
        
        # "../Bhb-17/Bhb-17-Darendeli_soft2",
        # "../Bhb-17/Bhb-17-Zhang_soft2",
        # 'DS_Template'
    ]
    with Pool(8) as pool:
            pool.map(generate_one_profile, profiles)

if __name__ == "__main__":
    # generate_one_profile('Bhd-01-Simplified_new')
    parallel_generate()
